import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def modifikace_inputexcelfile(file_path):
    # Načtení dat do pandas DataFrame
    df = pd.read_excel(file_path)

    # Načtení sešitu pomocí openpyxl
    wb = load_workbook(file_path)
    ws = wb.active

    # Nastavení stylů
    bold_font = Font(bold=True)
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Formátování hlaviček (první řádek) a modré pozadí pouze pro první řádek
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = blue_fill

    # Přidání tenké mřížky okolo všech buněk
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Automatické přizpůsobení šířky sloupců podle obsahu
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Uložení upraveného souboru
    wb.save(file_path)

# Testování funkce
if __name__ == "__main__":
    file_path = r'C:\Users\elala\Documents\UiPath\Process\InputFolder\RČ_IČ.xlsx'
    modifikace_inputexcelfile(file_path)
